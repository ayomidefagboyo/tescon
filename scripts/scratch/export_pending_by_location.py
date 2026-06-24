#!/usr/bin/env python3
"""
export_pending_by_location.py
=============================
Export the symbols that are GENUINELY NOT DONE, arranged by Location, in the
clean "uncaptured" style:  S/N | Symbol Number | Location | Desc1 | Desc2 | BOH

Truth comes from R2 itself (scanned LIVE every run), not from the dashboard's
cached tracker DB:

  • DONE  / CAPTURED -> a folder exists at  parts/{symbol}/   (final processed
                       images — this is the only thing that means "done in R2")
  • QUEUED / STAGED  -> raw images sitting at raw/{symbol}/ or listed in a
                       jobs/queued/*.json, but NOT yet processed into parts/
  • PENDING          -> in the catalog, but NEITHER done NOR queued

The "Pending by Location" sheet excludes BOTH done and queued. Queued items are
only staged (not finished), so they are kept on a SEPARATE sheet and are never
counted as pending — exactly as requested.

Why this differs from the tracking dashboard's number:
  - This script reads parts/ straight from R2, so "done" = what is really there.
  - The dashboard reads a cached tracker that (a) only matches R2 after a
    /tracker/sync-from-r2, and (b) computes "remaining" as
    total(21,780) - |parts/| - |raw-only|, subtracting EVERY parts/ folder even
    those whose symbol is not in the catalog. That makes its headline a derived
    count, not a clean list. This export gives the clean list instead.

Run on the Mac (needs R2 creds in backend/.env):
    python3 backend/export_pending_by_location.py
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path

import boto3
from botocore.config import Config
from dotenv import load_dotenv
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_ROOT.parent
load_dotenv(BACKEND_ROOT / ".env")

MASTER_CATALOG = BACKEND_ROOT / "Total EGTL Photo Project.xlsx"
OUT_DIR = PROJECT_ROOT / "reports"


def make_r2_client():
    acct = os.getenv("CLOUDFLARE_ACCOUNT_ID")
    ak = os.getenv("CLOUDFLARE_ACCESS_KEY_ID")
    sk = os.getenv("CLOUDFLARE_SECRET_ACCESS_KEY")
    bucket = os.getenv("CLOUDFLARE_BUCKET_NAME", "tescon-images")
    if not all([acct, ak, sk]):
        raise RuntimeError("Missing R2 credentials in backend/.env")
    client = boto3.client(
        "s3",
        endpoint_url=f"https://{acct}.r2.cloudflarestorage.com",
        aws_access_key_id=ak,
        aws_secret_access_key=sk,
        region_name="auto",
        config=Config(connect_timeout=10, read_timeout=30,
                      retries={"max_attempts": 3}, signature_version="s3v4"),
    )
    return client, bucket


def norm(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()


def loose(v) -> str:
    """Aggressive normalization for catching format mismatches (report-only)."""
    return re.sub(r"[^A-Z0-9]", "", norm(v).upper())


def scan_folders(r2, bucket: str, prefix: str) -> set[str]:
    """One entry per immediate sub-folder name under `prefix` (fast)."""
    paginator = r2.get_paginator("list_objects_v2")
    out = set()
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix, Delimiter="/"):
        for pre in page.get("CommonPrefixes", []):
            sym = pre["Prefix"].split("/")[1]
            if sym:
                out.add(norm(sym))
    return out


def scan_queued_jobs(r2, bucket: str) -> set[str]:
    """Symbols listed inside any jobs/queued/*.json manifest."""
    paginator = r2.get_paginator("list_objects_v2")
    out = set()
    for page in paginator.paginate(Bucket=bucket, Prefix="jobs/queued/"):
        for obj in page.get("Contents", []):
            if not obj["Key"].endswith(".json"):
                continue
            try:
                data = json.loads(r2.get_object(Bucket=bucket, Key=obj["Key"])["Body"].read())
                for p in data.get("parts", []):
                    if p.get("symbol_number"):
                        out.add(norm(p["symbol_number"]))
            except Exception as e:
                print(f"  (skip {obj['Key']}: {e})")
    return out


def style_sheet(ws, frame):
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    for c in range(1, frame.shape[1] + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    if frame.shape[0] > 0:
        ws.auto_filter.ref = ws.dimensions
    for c in range(1, frame.shape[1] + 1):
        col = frame.columns[c - 1]
        lengths = frame[col].astype(str).str.len().head(2000).tolist() or [0]
        width = max(len(str(col)), *lengths)
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 55)


def clean_frame(rows: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame({
        "Symbol Number": rows["Symbol Number"].apply(norm),
        "Location":      rows["Location"].apply(norm),
        "Desc1":         rows["Desc1"],
        "Desc2":         rows["Desc2"],
        "BOH":           rows["BOH"],
    })
    out = out.sort_values(["Location", "Symbol Number"], kind="stable").reset_index(drop=True)
    out.insert(0, "S/N", range(1, len(out) + 1))
    return out


def main() -> int:
    r2, bucket = make_r2_client()
    r2.list_objects_v2(Bucket=bucket, Prefix="parts/", MaxKeys=1)  # connectivity probe
    print(f"Connected to bucket: {bucket}\n")

    # --- Ground truth from R2 ------------------------------------------------
    print("Scanning R2 parts/ (DONE) …")
    captured = scan_folders(r2, bucket, "parts/")
    print(f"  → {len(captured):,} symbols actually done in R2 (parts/)")

    print("Scanning R2 raw/ + jobs/queued/ (STAGED) …")
    staged = (scan_folders(r2, bucket, "raw/") | scan_queued_jobs(r2, bucket)) - captured
    print(f"  → {len(staged):,} symbols staged but not done (excluded from pending)")

    # --- Catalog -------------------------------------------------------------
    print(f"\nReading catalog: {MASTER_CATALOG.name}")
    df = pd.read_excel(MASTER_CATALOG, sheet_name="Photo Data")
    df["sym_norm"] = df["Symbol Number"].apply(norm)
    cat_syms = set(df["sym_norm"])

    not_done = df[~df["sym_norm"].isin(captured)].copy()
    is_queued = not_done["sym_norm"].isin(staged)
    pending_rows = not_done[~is_queued]
    queued_rows = not_done[is_queued]

    pending = clean_frame(pending_rows)
    queued = clean_frame(queued_rows)

    # Per-location summary (pending counts only — queued is excluded entirely)
    summary = (
        pending.groupby("Location").size()
        .reset_index(name="Pending Count")
        .sort_values("Location", kind="stable").reset_index(drop=True)
    )

    # --- Write workbook (pending only — no queued sheet) ---------------------
    OUT_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    out_file = OUT_DIR / f"Pending_By_Location_{stamp}.xlsx"
    with pd.ExcelWriter(out_file, engine="openpyxl") as xw:
        pending.to_excel(xw, index=False, sheet_name="Pending by Location")
        summary.to_excel(xw, index=False, sheet_name="Location Summary")
        style_sheet(xw.sheets["Pending by Location"], pending)
        style_sheet(xw.sheets["Location Summary"], summary)

    # --- Reconciliation + "what's actually done in R2" diagnostic ------------
    matched = captured & cat_syms
    unmatched = sorted(captured - cat_syms)  # parts/ folders not in the catalog

    # Of those unmatched folders, how many look like a catalog symbol that we are
    # currently calling "pending" (i.e. a format mismatch hiding a done item)?
    pend_syms = set(pending["Symbol Number"])
    cat_loose = {}
    for s in cat_syms:
        cat_loose.setdefault(loose(s), []).append(s)
    likely_done_but_pending = []
    for folder in unmatched:
        hits = [s for s in cat_loose.get(loose(folder), []) if s in pend_syms]
        for h in hits:
            likely_done_but_pending.append((folder, h))

    print("\n" + "=" * 64)
    print("RECONCILIATION  (everything below is live from R2)")
    print("=" * 64)
    print(f"  Catalog symbols (Photo Data):        {len(df):,}")
    print(f"  Done in R2  (parts/ folders):        {len(captured):,}")
    print(f"     ...matched to a catalog symbol:   {len(matched):,}")
    print(f"     ...NOT in the catalog:            {len(unmatched):,}")
    print(f"  Staged not done (raw/ or queued):    {len(staged):,}")
    print("  " + "-" * 50)
    print(f"  PENDING  (exported, by location):    {len(pending):,}")
    print(f"  QUEUED   (excluded from pending):     {len(queued):,}")
    print(f"  Pending + Queued = not-done catalog: {len(pending) + len(queued):,}")

    if unmatched:
        print(f"\n  ⚠ {len(unmatched):,} parts/ folders are NOT in the catalog. Examples:")
        print("    " + ", ".join(unmatched[:10]) + (" ..." if len(unmatched) > 10 else ""))
    if likely_done_but_pending:
        print(f"\n  ⚠ {len(likely_done_but_pending):,} of those look like a FORMAT-MISMATCHED")
        print("    catalog symbol that is currently being listed as PENDING")
        print("    (i.e. likely already done in R2, just stored with different formatting):")
        for folder, sym in likely_done_but_pending[:10]:
            print(f"      R2 folder '{folder}'  ≈  catalog '{sym}'")
        print("    → tell me if you want these treated as done and removed from pending.")

    print(f"\n✅ Wrote {len(pending):,} pending symbols across "
          f"{len(summary):,} locations")
    print(f"   {out_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
