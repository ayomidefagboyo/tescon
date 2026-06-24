import os
import sys
from pathlib import Path
import pandas as pd
from datetime import datetime, timedelta
from dotenv import load_dotenv

# Load environment variables
backend_root = Path(__file__).resolve().parent
load_dotenv(backend_root / ".env")

if str(backend_root) not in sys.path:
    sys.path.insert(0, str(backend_root))

from app.services.cloudflare_r2 import get_r2_storage
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_boss_report():
    print("Connecting to live R2 storage to get 100% accurate numbers...")
    
    r2 = get_r2_storage()
    if not r2:
        print("Error: Could not connect to R2 storage. Please check your .env credentials.")
        return
        
    paginator = r2.s3_client.get_paginator('list_objects_v2')
    processed_parts = set()
    total_images_in_r2 = 0
    
    earliest_date = None
    latest_date = None
    
    print("Scanning R2 'parts/' directory...")
    # Using PageIterator to go through all objects in the bucket
    for page in paginator.paginate(Bucket=r2.bucket_name, Prefix='parts/'):
        for obj in page.get('Contents', []):
            total_images_in_r2 += 1
            key = obj['Key']
            last_mod = obj['LastModified'].replace(tzinfo=None)
            
            # Extract symbol number from path: parts/SYMBOL_NUMBER/filename
            paths = key.split('/')
            if len(paths) >= 2:
                processed_parts.add(paths[1])
                
            if earliest_date is None or last_mod < earliest_date:
                earliest_date = last_mod
            if latest_date is None or last_mod > latest_date:
                latest_date = last_mod
                
    processed_count = len(processed_parts)
    print(f"Found {processed_count} unique tags (parts) and {total_images_in_r2} total images captured.")
    
    # 2. Load Master Catalog for total count and details
    catalog_path = str(backend_root / 'Total EGTL Photo Project.xlsx')
    try:
        print(f"Loading master catalog from {catalog_path}...")
        df = pd.read_excel(catalog_path, sheet_name='Photo Data')
    except Exception as e:
        print(f"Error loading master catalog: {e}")
        return
        
    total_tags = len(df)
    outstanding = total_tags - processed_count
    percentage = (processed_count / total_tags) * 100 if total_tags > 0 else 0
    
    # Estimate completion month
    completion_month = "TBD"
    if earliest_date and latest_date:
        days_elapsed = (latest_date - earliest_date).days + 1
        if days_elapsed > 0:
            daily_rate = processed_count / days_elapsed
            if daily_rate > 0:
                days_remaining = outstanding / daily_rate
                completion_date = datetime.now() + timedelta(days=days_remaining)
                completion_month = completion_date.strftime("%B %Y")
    
    print(f"Metrics: Total={total_tags}, Captured={processed_count}, Outstanding={outstanding}, Completion={completion_month}")
    
    downloads_dir = Path("/Users/admin/tescon/downloads")
    downloads_dir.mkdir(parents=True, exist_ok=True)
    output_file = str(downloads_dir / "EGTL_Photo_Project_Dashboard.xlsx")
    print(f"Creating Excel report: {output_file}...")
    
    # Create Workbook
    wb = Workbook()
    
    # ---------------------------------------------------------
    # Sheet 1: Dashboard
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Dashboard"
    
    # Title
    ws1['A1'] = "PROJECT DASHBOARD"
    ws1['A1'].font = Font(size=18, bold=True, color="003366")
    ws1.merge_cells('A1:C1')
    
    metrics = [
        ("TOTAL NUMBER OF TAGS", f"{total_tags:,}"),
        ("QTY OF IMAGES CAPTURED TO DATE", f"{processed_count:,}"),
        ("QTY OF IMAGES OUTSTANDING", f"{outstanding:,}"),
        ("PERCENTAGE COMPLETED", f"{percentage:.1f}%"),
        ("PROJECT COMPLETION MONTH", completion_month)
    ]
    
    # Write metrics starting at row 3
    for idx, (label, value) in enumerate(metrics, start=3):
        cell_label = ws1[f'A{idx}']
        cell_label.value = label
        cell_label.font = Font(bold=True)
        
        cell_value = ws1[f'B{idx}']
        cell_value.value = value
        cell_value.font = Font(bold=True)
        cell_value.alignment = Alignment(horizontal="right")
        
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 20
    
    # Hide the data for the chart below
    data_start_row = 15
    ws1[f'A{data_start_row}'] = "Status"
    ws1[f'B{data_start_row}'] = "Count"
    ws1[f'A{data_start_row+1}'] = "Captured"
    ws1[f'B{data_start_row+1}'] = processed_count
    ws1[f'A{data_start_row+2}'] = "Outstanding"
    ws1[f'B{data_start_row+2}'] = outstanding
    
    # Create Pie Chart
    pie = PieChart()
    pie.title = "Project Completion Status"
    
    labels = Reference(ws1, min_col=1, min_row=data_start_row+1, max_row=data_start_row+2)
    data = Reference(ws1, min_col=2, min_row=data_start_row, max_row=data_start_row+2)
    
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    
    # Show percentage on data labels
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    
    ws1.add_chart(pie, "D3")
    
    # ---------------------------------------------------------
    # Prepare Dataframes
    # ---------------------------------------------------------
    df['Symbol Number'] = df['Symbol Number'].astype(str)
    processed_set = set(str(p) for p in processed_parts)
    
    # Add Status column
    df['Capture Status'] = df['Symbol Number'].apply(lambda x: 'Captured' if x in processed_set else 'Outstanding')
    
    df_captured = df[df['Capture Status'] == 'Captured']
    df_outstanding = df[df['Capture Status'] == 'Outstanding']
    
    def add_sheet_from_df(wb, df, sheet_title):
        print(f"Preparing '{sheet_title}' sheet...")
        ws = wb.create_sheet(title=sheet_title)
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            # Only check first 100 rows for length to speed up
            for cell in list(col)[:100]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column].width = adjusted_width

    # ---------------------------------------------------------
    # Add Data Sheets
    # ---------------------------------------------------------
    add_sheet_from_df(wb, df_captured, "Captured Parts")
    add_sheet_from_df(wb, df_outstanding, "Outstanding Parts")
    add_sheet_from_df(wb, df, "All Parts")

    wb.save(output_file)
    print(f"✅ Report successfully generated: {output_file}")

if __name__ == "__main__":
    generate_boss_report()
