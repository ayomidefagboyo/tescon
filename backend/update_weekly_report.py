"""
Refresh the EGTL Photo Project Weekly Report with the latest data from
Cloudflare R2.

Source of truth:
    1. Cloudflare R2 — gives us the live set of Symbol Numbers that have been
       photographed (one folder per part under `parts/`).
    2. Total EGTL Photo Project.xlsx (Photo Data sheet) — master catalog of
       every Symbol Number on the project (21,780 tags).

What this script updates in the workbook:
    * `Data` sheet            — replaced with the rows from the master catalog
                                 whose Symbol Number is present in R2.
    * `Summary Data` sheet    — refreshes Captured / Outstanding / Percentage
                                 / Pie Chart Data values. The Dashboard pie
                                 chart references these cells, so it updates
                                 automatically.

Output:
    Saves a new revision next to the original, named
        EGTL_Photo_Project_Weekly Report (<Today>) Rev <N>.xlsx
    so the previous revision is preserved.

Run:
    python3 backend/update_weekly_report.py
"""
from __future__ import annotations

import os
import re
import sys
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
import argparse

BACKEND_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_ROOT.parent

# Load env (R2 credentials live in backend/.env)
load_dotenv(BACKEND_ROOT / ".env")

# Make sibling `app` package importable so we can reuse the existing R2 client.
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from app.services.cloudflare_r2 import get_r2_storage  # noqa: E402

# -----------------------------------------------------------------------------
# Config
# -----------------------------------------------------------------------------
TEMPLATE_FILE = PROJECT_ROOT / "archive" / "reports" / "EGTL_Photo_Project_Weekly Report (June 9th, 2026).xlsx"
MASTER_CATALOG = BACKEND_ROOT / "data" / "Total EGTL Photo Project.xlsx"
OUTPUT_DIR = PROJECT_ROOT  # save next to the template

DATA_SHEET = "Data"
SUMMARY_SHEET = "Summary Data"

# Cells in `Summary Data` that we refresh. Layout was confirmed against the
# May 05 2026 Rev 1 file:
#   B4 Total Number of Tags      | C4 value
#   B5 QTY Captured to Date      | C5 value
#   B6 QTY Outstanding           | C6 value
#   B7 Percentage Completed      | C7 value (decimal, formatted as %)
#   B8 Captured Symbol Rows      | C8 value
#   B14 Completed (pie data)     | C14 quantity   D14 share
#   B15 Outstanding (pie data)   | C15 quantity   D15 share
SUMMARY_CELLS = {
    "total":           "C4",
    "captured":        "C5",
    "outstanding":     "C6",
    "percentage":      "C7",
    "captured_rows":   "C8",
    "pie_completed":   "C14",
    "pie_completed_p": "D14",
    "pie_outstanding": "C15",
    "pie_outstanding_p": "D15",
}


# -----------------------------------------------------------------------------
# Step 1 — Pull captured Symbol Numbers from R2
# -----------------------------------------------------------------------------
def fetch_captured_symbols_from_r2(as_of: datetime = None) -> tuple[set[str], int]:
    """Return (set of symbol numbers, total image count) found in R2.

    Uses Delimiter='/' to list only the symbol-number subdirectory prefixes
    under parts/ rather than iterating every individual image file.  This is
    ~10-50x faster because the bucket contains ~45 k images but only ~15 k
    symbol folders.

    When --as-of is supplied we fall back to full object listing so we can
    filter by LastModified date.
    """
    print("Connecting to Cloudflare R2…")
    r2 = get_r2_storage()
    if r2 is None:
        raise RuntimeError(
            "Cloudflare R2 not available. Check CLOUDFLARE_* env vars in backend/.env."
        )

    paginator = r2.s3_client.get_paginator("list_objects_v2")
    captured: set[str] = set()

    if as_of is not None:
        # Need per-object LastModified — full scan required
        print(f"Scanning all objects under 'parts/' (as-of filter active)…")
        image_count = 0
        for page in paginator.paginate(Bucket=r2.bucket_name, Prefix="parts/"):
            for obj in page.get("Contents", []):
                if obj["LastModified"] > as_of:
                    continue
                image_count += 1
                parts = obj["Key"].split("/")
                if len(parts) >= 2 and parts[1]:
                    captured.add(parts[1])
    else:
        # Fast path: list folder prefixes only (Delimiter="/")
        print(f"Listing symbol folders under 'parts/' (fast prefix scan)…")
        for page in paginator.paginate(
            Bucket=r2.bucket_name, Prefix="parts/", Delimiter="/"
        ):
            for prefix in page.get("CommonPrefixes", []):
                # prefix looks like "parts/12345678/"
                sym = prefix["Prefix"].split("/")[1]
                if sym:
                    captured.add(sym)
        image_count = len(captured) * 3  # approximate; not needed for the report

    print(f"  → {len(captured):,} unique symbol numbers")
    return captured, image_count


# -----------------------------------------------------------------------------
# Step 2 — Load master catalog rows for the captured symbols
# -----------------------------------------------------------------------------
def load_master_catalog() -> tuple[list[str], list[list]]:
    """Return (header_row, all_data_rows) from the master catalog Photo Data sheet."""
    if not MASTER_CATALOG.exists():
        raise FileNotFoundError(f"Master catalog not found: {MASTER_CATALOG}")

    print(f"Loading master catalog: {MASTER_CATALOG.name}")
    wb = openpyxl.load_workbook(MASTER_CATALOG, read_only=True, data_only=True)
    ws = wb["Photo Data"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Master catalog 'Photo Data' sheet is empty.")
    header, *data = rows
    print(f"  → {len(data):,} master catalog rows, {len(header)} columns")
    return list(header), [list(r) for r in data]


def normalize(value) -> str:
    """Normalize a Symbol Number for cross-file comparison (strings vs ints)."""
    if value is None:
        return ""
    s = str(value).strip()
    # 39018958 vs "39018958" — keep digits only when it's clearly numeric.
    return s


# -----------------------------------------------------------------------------
# Step 3 — Rewrite the workbook
# -----------------------------------------------------------------------------
def replace_data_sheet(wb: openpyxl.Workbook, header: list, rows: list[list]) -> None:
    """Replace the body of the Data sheet, preserving header style."""
    ws = wb[DATA_SHEET]

    # Capture existing header style so we don't lose formatting.
    header_styles = []
    for col_idx in range(1, ws.max_column + 1):
        cell: Cell = ws.cell(row=1, column=col_idx)
        header_styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "number_format": cell.number_format,
        })

    # Wipe everything past the header.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Rewrite header (in case column count changed) — ensure all columns present.
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        if col_idx - 1 < len(header_styles):
            s = header_styles[col_idx - 1]
            cell.font = s["font"]
            cell.fill = s["fill"]
            cell.alignment = s["alignment"]
            cell.border = s["border"]
            cell.number_format = s["number_format"]

    # Write data rows.
    for row in rows:
        ws.append(row)

    # Auto-size columns (cap at 40) to keep things readable.
    for col_idx in range(1, len(header) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(header[col_idx - 1] or ""))
        for row in rows[:200]:  # sample first 200 rows for width
            v = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    print(f"  → Data sheet now contains {len(rows):,} captured rows")


def update_summary_sheet(
    wb: openpyxl.Workbook,
    total: int,
    captured: int,
    outstanding: int,
) -> None:
    ws = wb[SUMMARY_SHEET]
    pct = (captured / total) if total else 0

    ws[SUMMARY_CELLS["total"]] = total
    ws[SUMMARY_CELLS["captured"]] = captured
    ws[SUMMARY_CELLS["outstanding"]] = outstanding
    ws[SUMMARY_CELLS["percentage"]] = pct
    ws[SUMMARY_CELLS["captured_rows"]] = captured

    ws[SUMMARY_CELLS["pie_completed"]] = captured
    ws[SUMMARY_CELLS["pie_completed_p"]] = pct
    ws[SUMMARY_CELLS["pie_outstanding"]] = outstanding
    ws[SUMMARY_CELLS["pie_outstanding_p"]] = 1 - pct if total else 0

    # Make sure percentage cells render as percentages.
    ws[SUMMARY_CELLS["percentage"]].number_format = "0.00%"
    ws[SUMMARY_CELLS["pie_completed_p"]].number_format = "0.00%"
    ws[SUMMARY_CELLS["pie_outstanding_p"]].number_format = "0.00%"

    print(
        f"  → Summary updated: total={total:,} captured={captured:,} "
        f"outstanding={outstanding:,} ({pct*100:.2f}%)"
    )

def update_dashboard_sheet(wb: openpyxl.Workbook, total: int, captured: int) -> None:
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        outstanding = total - captured
        pct = (captured / total) if total else 0
        
        ws["C7"] = total
        ws["E7"] = captured
        ws["G7"] = outstanding
        ws["I7"] = pct
        
        ws["C11"] = f"{pct * 100:.1f}% completed — {captured:,} captured / {total:,} total tags"
        print(f"  → Dashboard sheet updated: {ws['C11'].value}")


# -----------------------------------------------------------------------------
# Step 4 — Output filename (auto-bump the Rev counter)
# -----------------------------------------------------------------------------
def build_output_path(as_of: datetime = None) -> Path:
    dt = as_of if as_of else datetime.now()
    today = dt.strftime("%b %d, %Y")  # e.g. "May 10, 2026"
    base = f"EGTL_Photo_Project_Weekly Report ({today})"

    existing_revs = []
    pattern = re.compile(re.escape(base) + r" Rev (\d+)\.xlsx$")
    for p in OUTPUT_DIR.glob(f"{base} Rev *.xlsx"):
        m = pattern.search(p.name)
        if m:
            existing_revs.append(int(m.group(1)))

    next_rev = (max(existing_revs) + 1) if existing_revs else 1
    return OUTPUT_DIR / f"{base} Rev {next_rev}.xlsx"

# Incremental helper — load existing captured rows from a base report
# -----------------------------------------------------------------------------
def load_base_report_rows(base_path: str) -> tuple[list, list[list], set[str]]:
    """
    Read the Data sheet from a previous report.
    Returns (header, existing_data_rows, set_of_normalised_symbol_numbers).
    """
    print(f"Reading base report: {Path(base_path).name}")
    wb = openpyxl.load_workbook(base_path, read_only=True, data_only=True)
    ws = wb["Data"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], set()
    header = list(rows[0])
    data = [list(r) for r in rows[1:]]
    try:
        sym_idx = header.index("Symbol Number")
    except ValueError:
        sym_idx = None
    existing_syms = {normalize(r[sym_idx]) for r in data if sym_idx is not None and r[sym_idx]}
    print(f"  → {len(data):,} existing captured rows, {len(existing_syms):,} unique symbols")
    return header, data, existing_syms


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--as-of", help="Format: YYYY-MM-DD")
    parser.add_argument("--local-data-source", help="Bypass R2: read captured symbols from an existing report's Data sheet")
    parser.add_argument("--base-report", help=(
        "Incremental mode: path to the previous report. "
        "Existing rows are kept as-is; only NEW symbols found in R2 are looked up "
        "in the master catalog and appended. Much faster than a full rebuild."
    ))
    args = parser.parse_args()

    as_of_dt = None
    if args.as_of:
        as_of_dt = datetime.strptime(args.as_of, "%Y-%m-%d").replace(
            tzinfo=timezone.utc, hour=23, minute=59, second=59
        )

    if not TEMPLATE_FILE.exists():
        print(f"ERROR: template not found: {TEMPLATE_FILE}", file=sys.stderr)
        return 1

    # ------------------------------------------------------------------
    # Incremental mode
    # ------------------------------------------------------------------
    if args.base_report:
        base_header, base_rows, base_syms = load_base_report_rows(args.base_report)

        # Get full current set from R2 (fast prefix scan)
        r2_symbols, image_count = fetch_captured_symbols_from_r2(as_of_dt)
        r2_norm = {normalize(s) for s in r2_symbols}

        # Only the brand-new ones need a master-catalog lookup
        new_syms = r2_norm - base_syms
        print(f"\n  → {len(new_syms):,} new symbols since base report")

        if new_syms:
            cat_header, all_cat_rows = load_master_catalog()
            # Align columns: use master catalog header as the authoritative one
            try:
                sym_idx = cat_header.index("Symbol Number")
            except ValueError:
                print("ERROR: master catalog has no 'Symbol Number' column.", file=sys.stderr)
                return 1
            new_cat_rows = [r for r in all_cat_rows if normalize(r[sym_idx]) in new_syms]
            print(f"  → {len(new_cat_rows):,} new rows matched in master catalog")
            header = cat_header
            total_captured_rows = base_rows + new_cat_rows
        else:
            print("  → No new symbols — report numbers are unchanged from base.")
            cat_header, all_cat_rows = load_master_catalog()
            header = cat_header if base_header == cat_header else base_header
            total_captured_rows = base_rows

        total_catalog = len(all_cat_rows) if new_syms else 21780  # use known count if we skipped catalog load
        if new_syms:
            total_catalog = len(all_cat_rows)
        captured = len(total_captured_rows)
        outstanding = total_catalog - captured

        print(f"\nTotal captured: {captured:,} / {total_catalog:,} ({captured/total_catalog*100:.2f}%)")

    # ------------------------------------------------------------------
    # Local-data-source bypass (no R2, use an existing Data sheet as source of truth)
    # ------------------------------------------------------------------
    elif args.local_data_source:
        print(f"Bypassing R2, reading captured symbols from {args.local_data_source}...")
        source_wb = openpyxl.load_workbook(args.local_data_source, read_only=True, data_only=True)
        source_ws = source_wb["Data"]
        header_row = [c.value for c in source_ws[1]]
        try:
            sym_idx = header_row.index("Symbol Number")
        except ValueError:
            print("ERROR: Could not find Symbol Number in local data source")
            return 1
        captured_symbols = set()
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            if row[sym_idx]:
                captured_symbols.add(str(row[sym_idx]))
        image_count = len(captured_symbols) * 3

        cat_header, all_cat_rows = load_master_catalog()
        header = cat_header
        captured_norm = {normalize(s) for s in captured_symbols}
        total_captured_rows = [r for r in all_cat_rows if normalize(r[cat_header.index("Symbol Number")]) in captured_norm]
        total_catalog = len(all_cat_rows)
        captured = len(total_captured_rows)
        outstanding = total_catalog - captured

    # ------------------------------------------------------------------
    # Full rebuild from R2 (default)
    # ------------------------------------------------------------------
    else:
        captured_symbols, image_count = fetch_captured_symbols_from_r2(as_of_dt)
        if not captured_symbols:
            print("WARNING: no captured symbols found in R2 — aborting to protect existing report.")
            return 2
        cat_header, all_cat_rows = load_master_catalog()
        header = cat_header
        try:
            sym_idx = cat_header.index("Symbol Number")
        except ValueError:
            print("ERROR: master catalog has no 'Symbol Number' column.", file=sys.stderr)
            return 1
        captured_norm = {normalize(s) for s in captured_symbols}
        total_captured_rows = [r for r in all_cat_rows if normalize(r[sym_idx]) in captured_norm]
        total_catalog = len(all_cat_rows)
        captured = len(total_captured_rows)
        outstanding = total_catalog - captured
        print(f"\nMatched {captured:,} of {len(captured_symbols):,} R2 symbols against catalog "
              f"(total catalog rows: {total_catalog:,})")

    # ------------------------------------------------------------------
    # Write workbook
    # ------------------------------------------------------------------
    print(f"\nLoading workbook template: {TEMPLATE_FILE.name}")
    wb = openpyxl.load_workbook(TEMPLATE_FILE)

    print("Replacing Data sheet…")
    replace_data_sheet(wb, header, total_captured_rows)

    print("Refreshing Summary Data…")
    update_summary_sheet(wb, total=total_catalog, captured=captured, outstanding=outstanding)

    print("Updating Dashboard sheet…")
    update_dashboard_sheet(wb, total=total_catalog, captured=captured)

    output_path = build_output_path(as_of_dt)
    print(f"\nSaving updated workbook → {output_path.name}")
    wb.save(output_path)

    print("\n✅ Done.")
    print(f"   Captured symbols : {captured:,}")
    print(f"   Outstanding      : {outstanding:,}")
    print(f"   Output           : {output_path}")
    return 0



if __name__ == "__main__":
    sys.exit(main())
